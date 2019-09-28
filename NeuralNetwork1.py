# -*- coding: utf-8 -*-
"""
Created on Mon Jul  9 12:13:03 2018

@author: MHA
"""
import os
import numpy as np
import random
from sklearn.neural_network import MLPRegressor
from openpyxl import Workbook

def PercentageError(a,b):
    return (np.fabs(a-b)/a)*100

def NeuralNetworkAnalyse(X,y,X_Predicted,y_predicted,ws,no):
       
    mlp=MLPRegressor(hidden_layer_sizes=(4,),
                                         activation='tanh',
                                         solver='lbfgs',
                                         learning_rate='adaptive',
                                         max_iter=10000,
                                         learning_rate_init=0.01,
                                         alpha=3)
    
    mlp.fit(X,y)
    print('Fitting done!')
    y_pred=mlp.predict(X_Predicted)
    print('Actual y is : ')
    print(y_predicted)
    print('Neuaral Network Predicted :')
    print(y_pred)
    ws.cell(row=1,column=no*3+1).value="Actual Output"
    ws.cell(row=1,column=no*3+2).value="Predicted Output"
    ws.cell(row=1,column=no*3+3).value="Percentage Error"
    
    for i in range(len(y_pred)):
        ws.cell(row=i+2,column=no*3+1).value=y_predicted[i]
        ws.cell(row=i+2,column=no*3+2).value=y_pred[i]
        ws.cell(row=i+2,column=no*3+3).value=PercentageError(y_predicted[i],y_pred[i])
        
        

img_folder = os.getcwd()
img_files = [f for f in os.listdir(img_folder) if f.endswith('.mhzs')]

X = []
y = []

counter=0    #  No of instances
for f in img_files:    
    file=open(f,'r')
    myfiles=file.readlines()
    data=[]
    for j in myfiles:
        data.append(float(j))
    if float(data[0])>=4.5:
        y.append(data[0])
        counter+=1
        del data[0]      
        X.append(data)    
    file.close()
    
no_of_test_elements=int(counter/5)
print("No of test elements: "+str(no_of_test_elements))
wb=Workbook()
ws=wb.active
for i in range(10):      
    X_Pred=[]
    y_pred=[]
    X_Train=X.copy()
    y_train=y.copy()
    X_Test=[]
    print('Test no ... '+str(i+1))
    for j in range(no_of_test_elements):
        rnd=random.randint(0,counter-1)
        X_Pred.append(X_Train[rnd])
        y_pred.append(y_train[rnd])
        del X_Train[rnd]
        del y_train[rnd]
        counter-=1
    counter+=no_of_test_elements
    X_Train=np.array(X_Train,dtype=float)
    y_train=np.array(y_train,dtype=float)
    X_Pred=np.array(X_Pred,dtype=float)
    y_pred=np.array(y_pred,dtype=float)
    for k in range(no_of_test_elements):
        X_Test.append(X_Pred[k])
    NeuralNetworkAnalyse(X_Train,y_train,X_Test,y_pred,ws,i)
excel_path=os.path.join(os.getcwd(),"NNAnalyse.xlsx")
wb.save(excel_path)

#  Really Succeeded!!! Tnx ALLAH